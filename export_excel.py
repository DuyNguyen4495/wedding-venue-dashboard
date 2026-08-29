"""
Regenerates a styled Excel snapshot of dashboard/data/venues.csv.

Run from the dashboard/ folder:
    python export_excel.py

Output: ../Coyote_Hills_Venue_Tracker.xlsx (overwrites the existing tracker)
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

HERE = Path(__file__).parent
DATA_PATH = HERE / "data" / "venues.csv"
OUT_PATH = HERE.parent / "Coyote_Hills_Venue_Tracker.xlsx"

STATUS_FILLS = {
    "Baseline": "F3E6CF",
    "Sourced": "E7EBDF",
    "Not Sourced": "F2E2DD",
}


def fmt_money_range(low, high):
    if pd.isna(low) and pd.isna(high):
        return ""
    if pd.isna(high) or low == high:
        return f"${low:,.0f}"
    return f"${low:,.0f}-${high:,.0f}"


def build_display_frame(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    out["Status"] = df["status"]
    out["Venue Type"] = df["venue_type"]
    out["Venue"] = df["venue"]
    out["City"] = df["city"]
    out["Region"] = df["region"]
    out["Category"] = df["category"]
    out["Est. Total (100 gs.)"] = df.apply(lambda r: fmt_money_range(r["price_low"], r["price_high"]), axis=1)
    out["$ / Guest"] = df.apply(
        lambda r: fmt_money_range(r["price_per_guest_low"], r["price_per_guest_high"]), axis=1
    )
    out["Bar Included?"] = df["bar_included"].fillna("")
    out["Confidence"] = df["confidence"].fillna("")
    out["Notes / Next Step"] = df["notes"].fillna("")
    return out


def main():
    df = pd.read_csv(DATA_PATH)
    for col in ["price_low", "price_high", "price_per_guest_low", "price_per_guest_high"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    display = build_display_frame(df)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Venue Tracker"

    for row in dataframe_to_rows(display, index=False, header=True):
        ws.append(row)

    headers = list(display.columns)
    header_fill = PatternFill(start_color="3F5233", end_color="3F5233", fill_type="solid")
    header_font = Font(color="FFFDF9", bold=True, size=11)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    thin = Side(style="thin", color="DCD6C6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_col_idx = headers.index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col_idx).value
        fill_hex = STATUS_FILLS.get(status_val)
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid") if fill_hex else None
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 1:
                cell.font = Font(bold=True)

    widths = [12, 20, 42, 26, 22, 22, 22, 16, 24, 12, 62]
    for i, w in enumerate(widths, start=1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    last_col = get_column_letter(len(headers))
    last_row = ws.max_row
    tab = Table(displayName="VenueTracker", ref=f"A1:{last_col}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(tab)

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Rows: {len(df)}")


if __name__ == "__main__":
    main()
